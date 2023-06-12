import requests
from openpyxl import load_workbook
import datetime

# Materiales
materiales = {
    "power": 1,
    "transport": 13,
    "crude": 10,
    "metane": 74,
    "carbon_fiber": 75
}

# URLs
url = {
    "power": "https://www.simcompanies.com/api/v3/market/0/1/",
    "transport": "https://www.simcompanies.com/api/v3/market/0/13/",
    "crude": "https://www.simcompanies.com/api/v3/market/0/10/",
    "metane": "https://www.simcompanies.com/api/v3/market/0/74/",
    "carbon_fiber": "https://www.simcompanies.com/api/v3/market/0/75/"
}

# Proceso de información
def process_data(data):
    new_data = {}
    posted = data["posted"]
    datetime_obj = datetime.datetime.strptime(posted, "%Y-%m-%dT%H:%M:%S.%f%z")
    formatted_date = datetime_obj.strftime("%Y-%m-%d %H:%M:%S")
    new_data["posted"] = formatted_date
    new_data["seller"] = data["seller"]["company"]
    new_data["quantity"] = data["quantity"]
    new_data["quality"] = data["quality"]
    new_data["price"] = data["price"]
    return new_data


# Workbook
workbook = load_workbook(filename="Data.xlsx")
hojas = {
    "power": workbook["power"],
    "transport": workbook["transport"],
    "crude": workbook["crude"],
    "metane": workbook["metane"],
    "carbon_fiber": workbook["carbon_fiber"]
}

def encontrar_fila_vacia(hoja, calidad=None):
    fila_vacia = 3

    if calidad is not None:
        columna_inicio = (calidad * 5) + 1

        while hoja.cell(row=fila_vacia, column=columna_inicio).value is not None:
            fila_vacia += 1
        
        return fila_vacia
    else:
        columna_inicio = 1
        while hoja.cell(row=fila_vacia, column=columna_inicio).value is not None:
            fila_vacia += 1

        return fila_vacia


# Función para guardar los datos de calidad 0 en las columnas A-E
def guardar_datos_calidad_0(hoja, fila, datos):
    hoja.cell(row=fila, column=1).value = datos["posted"]
    hoja.cell(row=fila, column=2).value = datos["seller"]
    hoja.cell(row=fila, column=3).value = datos["quantity"]
    hoja.cell(row=fila, column=4).value = datos["quality"]
    hoja.cell(row=fila, column=5).value = datos["price"]

# Función para guardar los datos de calidad 1 en las columnas F-J
def guardar_datos_calidad_1(hoja, fila, datos):
    hoja.cell(row=fila, column=6).value = datos["posted"]
    hoja.cell(row=fila, column=7).value = datos["seller"]
    hoja.cell(row=fila, column=8).value = datos["quantity"]
    hoja.cell(row=fila, column=9).value = datos["quality"]
    hoja.cell(row=fila, column=10).value = datos["price"]

# Función para guardar los datos de calidad 2 en las columnas K-O
def guardar_datos_calidad_2(hoja, fila, datos):
    hoja.cell(row=fila, column=11).value = datos["posted"]
    hoja.cell(row=fila, column=12).value = datos["seller"]
    hoja.cell(row=fila, column=13).value = datos["quantity"]
    hoja.cell(row=fila, column=14).value = datos["quality"]
    hoja.cell(row=fila, column=15).value = datos["price"]

# Función para guardar los datos de calidad 3 en las columnas P-T
def guardar_datos_calidad_3(hoja, fila, datos):
    hoja.cell(row=fila, column=16).value = datos["posted"]
    hoja.cell(row=fila, column=17).value = datos["seller"]
    hoja.cell(row=fila, column=18).value = datos["quantity"]
    hoja.cell(row=fila, column=19).value = datos["quality"]
    hoja.cell(row=fila, column=20).value = datos["price"]

# Función para guardar los datos de calidad 4 en las columnas U-Y
def guardar_datos_calidad_4(hoja, fila, datos):
    hoja.cell(row=fila, column=21).value = datos["posted"]
    hoja.cell(row=fila, column=22).value = datos["seller"]
    hoja.cell(row=fila, column=23).value = datos["quantity"]
    hoja.cell(row=fila, column=24).value = datos["quality"]
    hoja.cell(row=fila, column=25).value = datos["price"]

registros_calidad = {
    'power': {0: 0, 1: 0, 2: 0, 3: 0, 4: 0},
    'transport': {0: 0, 1: 0, 2: 0, 3: 0, 4: 0},
    "crude": {0: 0, 1: 0, 2: 0, 3: 0, 4: 0},
    "metane": {0: 0, 1: 0, 2: 0, 3: 0, 4: 0},
    "carbon_fiber": {0: 0, 1: 0, 2: 0, 3: 0, 4: 0}
}


# Guardado de datos
for data in materiales.keys():
    response = requests.get(url[data])
    if response.status_code == 200:
        material_data = response.json()
        if data == "transport":
            hoja=hojas[data]
            for qly in material_data:
                quality = qly["quality"]
                if registros_calidad[data][quality] >= 10:
                    continue
                else:
                    fila_vacia = encontrar_fila_vacia(hoja,quality)
                    new_material_data = process_data(qly)
                    guardar_datos_calidad_0(hoja, fila_vacia, new_material_data)
                    registros_calidad[data][quality] += 1
        else:
            for qly in material_data:
                quality = qly["quality"]
                if quality not in registros_calidad[data]:
                    continue
                elif registros_calidad[data][quality] >= 10:
                    continue
                else:
                    new_material_data = process_data(qly)
                    hoja = hojas[data]
                    if quality == 0:
                        fila_vacia = encontrar_fila_vacia(hoja, quality)
                        guardar_datos_calidad_0(hoja, fila_vacia, new_material_data)
                    elif quality == 1:
                        fila_vacia = encontrar_fila_vacia(hoja, quality)
                        guardar_datos_calidad_1(hoja, fila_vacia, new_material_data)
                    elif quality == 2:
                        fila_vacia = encontrar_fila_vacia(hoja, quality)
                        guardar_datos_calidad_2(hoja, fila_vacia, new_material_data)
                    elif quality == 3:
                        fila_vacia = encontrar_fila_vacia(hoja, quality)
                        guardar_datos_calidad_3(hoja, fila_vacia, new_material_data)
                    elif quality == 4:
                        fila_vacia = encontrar_fila_vacia(hoja, quality)
                        guardar_datos_calidad_4(hoja, fila_vacia, new_material_data)
                registros_calidad[data][quality] += 1

    else:
        print(f"Error al obtener los datos de {data}, código de respuesta:", response.status_code)

workbook.save("Data.xlsx")
print(f'datos guardados\n{registros_calidad}')
